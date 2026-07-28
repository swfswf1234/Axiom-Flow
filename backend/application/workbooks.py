"""
模块职责：导出、导入并校验知识工作簿草稿和发布快照。
设计关联（DesignRef）：docs/design/excel-release-workflow.md
实现状态：Current
关联测试：tests/test_document_workflow.py
"""

import json
from pathlib import Path
from typing import Any, Protocol

from openpyxl import Workbook, load_workbook


class WorkbookRepository(Protocol):
    """工作簿用例所需的最小仓储端口。"""

    def accepted_snapshot(self, document_id: str) -> dict[str, Any]: ...
    def create_workbook_revision(
        self, document_id: str, path: Path, snapshot: dict[str, Any], status: str = "draft",
    ) -> dict[str, Any]: ...
    def latest_workbook_revision(self, document_id: str) -> dict[str, Any] | None: ...
    def create_release(self, document_id: str, revision_id: str, snapshot: dict[str, Any]) -> dict[str, Any]: ...


class WorkbookService:
    """工作簿只是人工编辑入口，发布前必须回到 MySQL 创建快照。"""

    def __init__(self, store: WorkbookRepository, data_dir: Path):
        self.store = store
        self.data_dir = data_dir

    def export_draft(self, document_id: str) -> dict[str, Any]:
        snapshot = self.store.accepted_snapshot(document_id)
        workbook = Workbook()
        workbook.remove(workbook.active)
        documents = workbook.create_sheet("documents")
        documents.append(["document_id"])
        documents.append([document_id])
        nodes = workbook.create_sheet("knowledge_nodes")
        nodes.append(["id", "kind", "title", "content", "evidence_json", "review_status"])
        for node in snapshot["nodes"]:
            nodes.append([node["id"], node["kind"], node["title"], node["content"], json.dumps(node["evidence"], ensure_ascii=False), node["review_status"]])
        edges = workbook.create_sheet("knowledge_edges")
        edges.append(["id", "source_id", "target_id", "relation", "evidence_json", "review_status"])
        for edge in snapshot["edges"]:
            edges.append([edge["id"], edge["source_id"], edge["target_id"], edge["relation"], json.dumps(edge["evidence"], ensure_ascii=False), edge["review_status"]])
        workbook.create_sheet("sections").append(["page_no", "title"])
        workbook.create_sheet("review_notes").append(["target_id", "note"])
        destination = self.data_dir / "documents" / document_id / "workbooks" / "draft.xlsx"
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(destination)
        return self.store.create_workbook_revision(document_id, destination, snapshot)

    def import_draft(self, document_id: str, path: Path) -> dict[str, Any]:
        workbook = load_workbook(path, data_only=True)
        required = {"documents", "sections", "knowledge_nodes", "knowledge_edges", "review_notes"}
        if not required <= set(workbook.sheetnames):
            raise ValueError("工作簿缺少必需工作表")
        document_rows = list(workbook["documents"].iter_rows(min_row=2, values_only=True))
        if len(document_rows) != 1 or str(document_rows[0][0] or "") != document_id:
            raise ValueError("工作簿的 document_id 与导入目标不一致")
        nodes = []
        for row in workbook["knowledge_nodes"].iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            evidence = self._evidence(row[4])
            title = str(row[2] or "").strip()
            if not title:
                raise ValueError("知识节点 title 不能为空")
            status = str(row[5] or "accepted")
            if status not in {"accepted", "rejected", "needs_review"}:
                raise ValueError("知识节点 review_status 不合法")
            nodes.append({"id": str(row[0]), "kind": str(row[1] or "concept"), "title": title, "content": str(row[3] or ""), "evidence": evidence, "review_status": status})
        node_ids = {node["id"] for node in nodes}
        edges = []
        for row in workbook["knowledge_edges"].iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            if str(row[1]) not in node_ids or str(row[2]) not in node_ids:
                raise ValueError("工作簿关系引用了不存在的节点")
            relation = str(row[3] or "RELATED_TO")
            if relation not in {"CONTAINS", "PREREQUISITE_OF", "DEFINES", "PROVES", "USES", "ILLUSTRATES", "RELATED_TO"}:
                raise ValueError("知识关系 relation 不合法")
            status = str(row[5] or "accepted")
            if status not in {"accepted", "rejected", "needs_review"}:
                raise ValueError("知识关系 review_status 不合法")
            edges.append({"id": str(row[0]), "source_id": str(row[1]), "target_id": str(row[2]), "relation": relation, "evidence": self._evidence(row[4]), "review_status": status})
        stored = self.data_dir / "documents" / document_id / "workbooks" / f"import-{path.name}"
        stored.parent.mkdir(parents=True, exist_ok=True)
        stored.write_bytes(path.read_bytes())
        return self.store.create_workbook_revision(document_id, stored, {"nodes": nodes, "edges": edges})

    def publish_latest(self, document_id: str) -> dict[str, Any]:
        revision = self.store.latest_workbook_revision(document_id)
        if not revision:
            raise ValueError("没有可发布的工作簿草稿")
        nodes = revision["snapshot"]["nodes"]
        if not nodes:
            raise ValueError("工作簿没有可发布的知识节点")
        if any(node["review_status"] != "accepted" for node in nodes):
            raise ValueError("工作簿包含未接受的知识节点")
        node_ids = {node["id"] for node in nodes}
        edges = revision["snapshot"]["edges"]
        if any(edge["review_status"] != "accepted" for edge in edges):
            raise ValueError("工作簿包含未接受的知识关系")
        if any(edge["source_id"] not in node_ids or edge["target_id"] not in node_ids for edge in edges):
            raise ValueError("工作簿关系引用了不存在的节点")
        return self.store.create_release(document_id, revision["id"], revision["snapshot"])

    @staticmethod
    def _evidence(value: Any) -> list[dict[str, Any]]:
        """导入时强制保留可回到 PDF 页的证据，拒绝无来源的手工知识。"""
        try:
            evidence = json.loads(value or "[]")
        except json.JSONDecodeError as exc:
            raise ValueError("证据不是合法 JSON") from exc
        if not isinstance(evidence, list) or not evidence:
            raise ValueError("知识节点和关系必须保留至少一条证据")
        if any(not isinstance(item, dict) or not isinstance(item.get("page_no"), int) for item in evidence):
            raise ValueError("每条证据必须包含整数 page_no")
        return evidence
