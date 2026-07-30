"""
模块职责：使用 OpenPyXL 读写并校验知识审阅工作簿文件。
设计关联（DesignRef）：docs/design/excel-release-workflow.md
实现状态：Current
关联测试：tests/system/test_document_release_flow.py
"""

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


class OpenPyxlWorkbookGateway:
    def __init__(self, data_dir: Path) -> None:
        self.data_dir = data_dir

    def export(self, document_id: str, snapshot: dict[str, Any]) -> Path:
        workbook = Workbook()
        workbook.remove(workbook.active)
        documents = workbook.create_sheet("documents")
        documents.append(["document_id"])
        documents.append([document_id])
        nodes = workbook.create_sheet("knowledge_nodes")
        nodes.append(["id", "kind", "title", "content", "evidence_json", "review_status"])
        for node in snapshot["nodes"]:
            nodes.append([node["id"], node["kind"], node["title"], node["content"], json.dumps(node["evidence"], ensure_ascii=False), node["review_status"]])
        edges = workbook.create_sheet("knowledge_edges")
        edges.append(["id", "source_id", "target_id", "relation", "evidence_json", "review_status"])
        for edge in snapshot["edges"]:
            edges.append([edge["id"], edge["source_id"], edge["target_id"], edge["relation"], json.dumps(edge["evidence"], ensure_ascii=False), edge["review_status"]])
        workbook.create_sheet("sections").append(["page_no", "title"])
        workbook.create_sheet("review_notes").append(["target_id", "note"])
        destination = self.data_dir / "documents" / document_id / "workbooks" / "draft.xlsx"
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(destination)
        return destination

    def import_file(self, document_id: str, path: Path) -> tuple[Path, dict[str, Any]]:
        workbook = load_workbook(path, data_only=True)
        required = {"documents", "sections", "knowledge_nodes", "knowledge_edges", "review_notes"}
        if not required <= set(workbook.sheetnames):
            raise ValueError("工作簿缺少必需工作表")
        document_rows = list(workbook["documents"].iter_rows(min_row=2, values_only=True))
        if len(document_rows) != 1 or str(document_rows[0][0] or "") != document_id:
            raise ValueError("工作簿的 document_id 与导入目标不一致")
        nodes = self._nodes(workbook)
        edges = self._edges(workbook, {node["id"] for node in nodes})
        stored = self.data_dir / "documents" / document_id / "workbooks" / f"import-{path.name}"
        stored.parent.mkdir(parents=True, exist_ok=True)
        stored.write_bytes(path.read_bytes())
        return stored, {"nodes": nodes, "edges": edges}

    @classmethod
    def _nodes(cls, workbook: Any) -> list[dict[str, Any]]:
        nodes = []
        for row in workbook["knowledge_nodes"].iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            title = str(row[2] or "").strip()
            if not title:
                raise ValueError("知识节点 title 不能为空")
            status = str(row[5] or "accepted")
            if status not in {"accepted", "rejected", "needs_review"}:
                raise ValueError("知识节点 review_status 不合法")
            nodes.append({"id": str(row[0]), "kind": str(row[1] or "concept"), "title": title, "content": str(row[3] or ""), "evidence": cls._evidence(row[4]), "review_status": status})
        return nodes

    @classmethod
    def _edges(cls, workbook: Any, node_ids: set[str]) -> list[dict[str, Any]]:
        edges = []
        allowed = {"CONTAINS", "PREREQUISITE_OF", "DEFINES", "PROVES", "USES", "ILLUSTRATES", "RELATED_TO"}
        for row in workbook["knowledge_edges"].iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            if str(row[1]) not in node_ids or str(row[2]) not in node_ids:
                raise ValueError("工作簿关系引用了不存在的节点")
            relation = str(row[3] or "RELATED_TO")
            if relation not in allowed:
                raise ValueError("知识关系 relation 不合法")
            status = str(row[5] or "accepted")
            if status not in {"accepted", "rejected", "needs_review"}:
                raise ValueError("知识关系 review_status 不合法")
            edges.append({"id": str(row[0]), "source_id": str(row[1]), "target_id": str(row[2]), "relation": relation, "evidence": cls._evidence(row[4]), "review_status": status})
        return edges

    @staticmethod
    def _evidence(value: Any) -> list[dict[str, Any]]:
        try:
            evidence = json.loads(value or "[]")
        except json.JSONDecodeError as exc:
            raise ValueError("证据不是合法 JSON") from exc
        if not isinstance(evidence, list) or not evidence:
            raise ValueError("知识节点和关系必须保留至少一条证据")
        if any(not isinstance(item, dict) or not isinstance(item.get("page_no"), int) for item in evidence):
            raise ValueError("每条证据必须包含整数 page_no")
        return evidence
